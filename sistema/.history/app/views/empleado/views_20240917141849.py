from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.shortcuts import render
from django.core.exceptions import ValidationError
from app.models import Empleado
from app.forms import EmpleadoForm
from django.views.decorators.cache import never_cache
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from openpyxl.drawing.image import Image

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.contrib import messages
from app.models import *
from app.forms import *
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib.colors import Color
from django.views.decorators.cache import never_cache

@never_cache
@login_required
def empleado_list(request):
    if not request.user.has_perm('app.view_empleado'):
        return render(request, 'empleado/listar.html', {'has_permission': False})

    empleados = Empleado.objects.all()

    estado_filtro = request.GET.get('estado', None)
    buscar_nombre = request.GET.get('buscar_nombre', '')
    buscar_telefono = request.GET.get('buscar_telefono', '')

    if estado_filtro == 'activado':
        empleados = empleados.filter(activado=True)
    elif estado_filtro == 'inactivado':
        empleados = empleados.filter(activado=False)

    if buscar_nombre:
        empleados = empleados.filter(nombre__icontains=buscar_nombre)

    if buscar_telefono:
        empleados = empleados.filter(telefono__icontains=buscar_telefono)

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Empleados', 'url': '/app/empleado/listar/'},
    ]

    context = {
        'titulo': 'Listado de empleados',
        'entidad': 'Empleados',
        'listar_url': reverse_lazy('empleado_lista'),
        'crear_url': reverse_lazy('empleado_crear'),
        'empleados': empleados,
        'breadcrumbs': breadcrumbs,
        'has_permission': request.user.has_perm('app.view_empleado'),
        'can_add': request.user.has_perm('app.add_empleado') and not request.user.groups.filter(name='Empleado').exists(),
    }
    return render(request, 'empleado/index.html', context)



@never_cache
@login_required
def empleado_create(request):
    if not request.user.has_perm('app.add_empleado') or request.user.groups.filter(name='Empleado').exists():
        return redirect('empleado_lista')

    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, '¡Empleado creado con éxito!')
                return redirect('empleado_lista')
            except ValidationError as e:
                form.add_error(None, e)
    else:
        form = EmpleadoForm()

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Empleados', 'url': '/app/empleado/listar/'},
        {'name': 'Registrar Empleado', 'url': '/app/empleado/crear/'},
    ]

    context = {
        'titulo': 'Registrar empleado',
        'entidad': 'Registrar empleado',
        'listar_url': reverse_lazy('empleado_lista'),
        'form': form,
        'breadcrumbs': breadcrumbs,
        'has_permission': not request.user.groups.filter(name='Empleado').exists() and request.user.has_perm('app.add_empleado'),
    }
    return render(request, 'empleado/crear.html', context)


@never_cache
@login_required
def empleado_update(request, id):
    if not request.user.has_perm('app.change_empleado') or request.user.groups.filter(name='Empleado').exists():
        return redirect('empleado_lista')

    empleado = get_object_or_404(Empleado, id=id)

    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleado)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, '¡Empleado editado con éxito!')
                return redirect('empleado_lista')
            except ValidationError as e:
                form.add_error(None, e)
    else:
        form = EmpleadoForm(instance=empleado)

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Empleados', 'url': '/app/empleado/listar/'},
        {'name': 'Editar Empleado', 'url': f'/app/empleado/editar/{id}/'},
    ]

    context = {
        'titulo': 'Editar empleado',
        'entidad': 'Empleado',
        'listar_url': reverse_lazy('empleado_lista'),
        'form': form,
        'breadcrumbs': breadcrumbs,
        'has_permission': request.user.has_perm('app.change_empleado') and not request.user.groups.filter(name='Empleado').exists(),
    }

    return render(request, 'empleado/crear.html', context)

@never_cache
@login_required
def empleado_delete(request, id):
    if not request.user.has_perm('app.delete_empleado') or request.user.groups.filter(name='Empleado').exists():
        return redirect('empleado_lista')

    empleado = get_object_or_404(Empleado, id=id)
    empleado.delete()
    messages.success(request, '¡Empleado eliminado con éxito!')
    return redirect('empleado_lista')


@never_cache
@login_required
def activar_empleado(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    empleado.activado = True
    empleado.save()
    return redirect('empleado_lista')


@never_cache
@login_required
def desactivar_empleado(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    empleado.activado = False
    empleado.save()
    return redirect('empleado_lista')



def reporte_empleados_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:F1')
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = "Empleados"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre", "Nombre de Usuario", "Email", "Tipo de Documento", "Número de Documento", "Teléfono"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    empleados = Empleado.objects.all()
    for empleado in empleados:
        ws.append([
            empleado.id,
            empleado.nombre,
            empleado.user.username,
            empleado.user.email,
            empleado.tipo_documento,
            empleado.numero_documento,
            empleado.telefono
        ])

    column_widths = [10, 30, 20, 30, 20, 20, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_empleados.xlsx"'
    
    wb.save(response)
    return response


from io import BytesIO
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Table, TableStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import Color
from app.models import Empleado

def reporte_empleados_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    extra_margin = 20
    row_height = 18

    watermark_path = finders.find('img/logo.png')
    p.saveState()

    p.setFillColor(Color(1, 1, 1, alpha=0.3))
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3))

    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')
    p.restoreState()

    y_position = height - margin
    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Empleados")
    y_position -= 50

    column_widths = [20, 100, 100, 150, 100, 80, 80]
    headers = ["ID", "Nombre", "Nombre de Usuario", "Email", "Tipo de Documento", "Número de Documento", "Teléfono"]
    data = [headers]

    empleados = Empleado.objects.all()
    for empleado in empleados:
        data.append([
            str(empleado.id),
            empleado.nombre,
            empleado.user.username,
            empleado.user.email,
            empleado.tipo_documento,
            empleado.numero_documento,
            empleado.telefono
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 8),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)

    table_width = sum(column_widths)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height - extra_margin

    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_empleados.pdf"'
    return response
