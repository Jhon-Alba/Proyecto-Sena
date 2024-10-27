from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.contrib import messages
from app.models import *
from app.forms import *
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib.colors import Color
from django.views.decorators.cache import never_cache


@never_cache
@login_required
def administrador_list(request):
    if not request.user.has_perm('app.view_administrador'):
        return render(request, 'administrador/index.html', {'has_permission': False})

    administradores = Administrador.objects.all()

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Administradores', 'url': '/app/administrador/'},
    ]

    context = {
        'titulo': 'Listado de administradores',
        'entidad': 'Administradores',
        'listar_url': reverse_lazy('administrador_lista'),
        'crear_url': reverse_lazy('administrador_crear'),
        'administradores': administradores,
        'breadcrumbs': breadcrumbs,
        'has_permission': request.user.has_perm('app.view_administrador'),
        'can_add': request.user.has_perm('app.add_administrador') and not request.user.groups.filter(name='Empleado').exists(),
    }
    return render(request, 'administrador/index.html', context)


@never_cache
@login_required
def administrador_create(request):
    if not request.user.has_perm('app.add_administrador') or request.user.groups.filter(name='Empleado').exists():
        return redirect('administrador_lista')

    if request.method == 'POST':
        form = AdministradorForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, '¡Administrador creado con éxito!')
                return redirect('administrador_lista')
            except ValidationError as e:
                form.add_error(None, e)
    else:
        form = AdministradorForm()

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Administradores', 'url': '/app/administrador/'},
        {'name': 'Registrar Administrador', 'url': '/app/administrador/crear/'},
    ]

    context = {
        'titulo': 'Registrar administrador',
        'entidad': 'Registrar administrador',
        'listar_url': reverse_lazy('administrador_lista'),
        'form': form,
        'breadcrumbs': breadcrumbs,
        'has_permission': not request.user.groups.filter(name='Empleado').exists() and request.user.has_perm('app.add_administrador'),
    }
    return render(request, 'administrador/crear.html', context)

########################################## EDITAR ADMINISTRADOR ##################################################
@never_cache
@login_required
def administrador_update(request, id):
    if not request.user.has_perm('app.change_administrador') or request.user.groups.filter(name='Empleado').exists():
        return redirect('administrador_lista')

    administrador = get_object_or_404(Administrador, id=id)

    if request.method == 'POST':
        form = AdministradorForm(request.POST, instance=administrador)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, '¡Administrador editado con éxito!')
                return redirect('administrador_lista')
            except ValidationError as e:
                form.add_error(None, e)
    else:
        form = AdministradorForm(instance=administrador)

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Administradores', 'url': '/app/administrador/'},
        {'name': 'Editar Administrador', 'url': f'/app/administrador/editar/{id}/'},
    ]

    context = {
        'titulo': 'Editar Administrador',
        'entidad': 'Administrador',
        'listar_url': reverse_lazy('administrador_lista'),
        'form': form,
        'breadcrumbs': breadcrumbs,
        'has_permission': request.user.has_perm('app.change_administrador') and not request.user.groups.filter(name='Empleado').exists(),
    }

    return render(request, 'administrador/crear.html', context)



#FILTROS Administrador ###############################################################################################################

def administrador_lista(request):
    if not request.user.has_perm('app.view_administrador'):
        return render(request, 'administrador/index.html', {'has_permission': False})
    
    estado_filtro = request.GET.get('estado', None)
    buscar_nombre = request.GET.get('buscar_nombre', '')
    buscar_telefono = request.GET.get('buscar_celular', '')
    
    administradores = Administrador.objects.all()


    context = {
        'administradores': administradores,
        'estado_filtro': estado_filtro,
        'buscar_nombre': buscar_nombre,
        'buscar_telefono': buscar_telefono,
       
    }

    return render(request, 'administrador/index.html', context)


############################ ELIMINAR ADMINISTRADOR ###########################################################
@never_cache
@login_required
def administrador_delete(request, id):
    if not request.user.has_perm('app.delete_administrador') or request.user.groups.filter(name='Empleado').exists():
        return redirect('administrador_lista')

    administrador = get_object_or_404(Administrador, id=id)
    administrador.delete()
    messages.success(request, '¡Administrador eliminado con éxito!')
    return redirect('administrador_lista')


def activar_administrador(request, pk):
    administrador = get_object_or_404(Administrador, pk=pk)
    administrador.activado = True
    administrador.save()
    return redirect('administrador_lista')  

def desactivar_administrador(request, pk):
    administrador = get_object_or_404(Administrador, pk=pk)
    administrador.activado = False
    administrador.save()
    return redirect('administrador_lista')  

################## REPORTE EXCEL ADMINISTRADORES ###################################################################

def reporte_administradores_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Administradores"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:F1')
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = "Administradores"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre", "Nombre de Usuario", "Email", "Tipo de Documento", "Número de Documento", "Teléfono"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    administradores = Administrador.objects.all()
    for admin in administradores:
        ws.append([
            admin.id,
            admin.nombre,
            admin.user.username,
            admin.user.email,
            admin.tipo_documento,
            admin.numero_documento,
            admin.telefono
        ])

    column_widths = [10, 30, 20, 30, 20, 20, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_administradores.xlsx"'
    
    wb.save(response)
    return response

########################################### REPORTE PDF ###########################################################

def reporte_administradores_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    extra_margin = 20
    row_height = 18

    watermark_path = finders.find('img/logo.png')
    p.saveState()

    p.setFillColor(Color(1, 1, 1, alpha=0.3))
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3))

    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')
    p.restoreState()

    y_position = height - margin
    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Administradores")
    y_position -= 50

    # Reduce the column widths
    column_widths = [20, 100, 100, 150, 100, 80, 80]
    headers = ["ID", "Nombre", "Nombre de Usuario", "Email", "Tipo de Documento", "Número de Documento", "Teléfono"]
    data = [headers]

    administradores = Administrador.objects.all()
    for admin in administradores:
        data.append([
            str(admin.id),
            admin.nombre,
            admin.user.username,
            admin.user.email,
            admin.tipo_documento,
            admin.numero_documento,
            admin.telefono
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 8),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)

    table_width = sum(column_widths)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height - extra_margin

    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_administradores.pdf"'
    return response





