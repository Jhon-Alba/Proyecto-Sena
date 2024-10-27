from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import *
from .forms import *
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.staticfiles import finders
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Table, TableStyle
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from django.urls import reverse_lazy
from django.core.exceptions import ValidationError
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.shortcuts import redirect
from django.contrib.auth.views import LoginView
from django.contrib.auth import logout
from django.views.generic import RedirectView
from django.views.decorators.cache import never_cache
from django.views import View
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.views.decorators.http import require_POST
from django.conf import settings
from django.http import HttpResponse, Http404
from django.utils.http import urlsafe_base64_decode
import os
import subprocess
from datetime import datetime


@never_cache
@login_required
def gestionar_respaldos(request):
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    print(f"Ruta del directorio de respaldos: {backup_dir}")

    # Crear el directorio si no existe
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    backups = []
    try:
        for filename in os.listdir(backup_dir):
            if filename.endswith('.sql'):
                file_path = os.path.join(backup_dir, filename)
                created_at = datetime.fromtimestamp(os.path.getctime(file_path))
                size = os.path.getsize(file_path)
                backups.append({
                    'id': filename,
                    'filename': filename,
                    'created_at': created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'size': f"{size / 1024 / 1024:.2f} MB"
                })
    except FileNotFoundError:
        messages.error(request, "No se pudo encontrar el directorio de respaldos.")

    backups.sort(key=lambda x: x['created_at'], reverse=True)

    # Paginación
    paginator = Paginator(backups, 10)  # 10 respaldos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Respaldos', 'url': '/libreria/respaldos/'},
    ]

    return render(request, 'backup.html', {'page_obj': page_obj, 'breadcrumbs': breadcrumbs})

@method_decorator(never_cache, name='post')
class CrearRespaldoView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Obtener la configuración de la base de datos
            db_settings = settings.DATABASES['default']
            db_name = db_settings['NAME']
            db_user = db_settings['USER']
            db_password = db_settings['PASSWORD']
            db_host = db_settings['HOST']
            db_port = db_settings['PORT']

            # Crear el nombre del archivo de respaldo
            filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
            backup_dir = os.path.join(settings.BASE_DIR, 'backups')
            backup_path = os.path.join(backup_dir, filename)

            # Asegurarse de que el directorio de respaldos existe
            os.makedirs(backup_dir, exist_ok=True)

            # Usa la ruta completa a mysqldump.exe
            mysqldump_path = r"C:\xampp\mysql\bin\mysqldump.exe"
                
            # Comando para crear el respaldo
            command = (
                f"\"{mysqldump_path}\" -h {db_host} -P {db_port} -u {db_user} -p{db_password} "
                f"{db_name} > \"{backup_path}\""
            )

            # Ejecutar el comando y capturar la salida
            result = subprocess.run(command, shell=True, capture_output=True, text=True)
            if result.returncode != 0:
                messages.error(request, f"Error al crear el respaldo: {result.stderr}")
            else:
                messages.success(request, f"Respaldo creado exitosamente: {filename}")
        except Exception as e:
            messages.error(request, f"Error al crear el respaldo: {str(e)}")

        return redirect('backup')

@method_decorator(never_cache, name='post')
class RestaurarRespaldoView(View):
    def post(self, request, *args, **kwargs):
        # Obtener el nombre del archivo desde el campo oculto
        backup_id = request.POST.get('respaldo_id')

        if backup_id:
            try:
                # Obtener la configuración de la base de datos
                db_settings = settings.DATABASES['default']
                db_name = db_settings['NAME']
                db_user = db_settings['USER']
                db_password = db_settings['PASSWORD']
                db_host = db_settings['HOST']
                db_port = db_settings['PORT']

                # Construir la ruta del archivo de respaldo
                backup_dir = os.path.join(settings.BASE_DIR, 'backups')
                backup_path = os.path.join(backup_dir, backup_id)

                # Verificar que el archivo de respaldo existe
                if not os.path.exists(backup_path):
                    messages.error(request, "El archivo de respaldo especificado no existe.")
                    return redirect('backup')

                # Escapar la ruta del archivo para evitar problemas con espacios
                backup_path_escaped = f"\"{backup_path}\""

                # Usa la ruta completa a mysql.exe
                mysql_path = r"C:\xampp\mysql\bin\mysql.exe"

                # Comando para restaurar la base de datos
                command = (
                    f"\"{mysql_path}\" -h {db_host} -P {db_port} -u {db_user} -p{db_password} "
                    f"{db_name} < {backup_path_escaped}"
                )

                # Imprimir el comando para verificar
                print(f"Comando a ejecutar: {command}")

                # Ejecutar el comando y capturar la salida
                result = subprocess.run(command, shell=True, capture_output=True, text=True)

                if result.returncode != 0:
                    messages.error(request, f"Error al restaurar la base de datos: {result.stderr}")
                else:
                    messages.success(request, f"Base de datos restaurada desde {backup_id}")

            except subprocess.CalledProcessError as e:
                messages.error(request, f"Error al ejecutar el comando de restauración: {e}")
            except Exception as e:
                messages.error(request, f"Error inesperado al restaurar la base de datos: {str(e)}")
        else:
            messages.error(request, "No se especificó un archivo para restaurar")

        return redirect('backup')

@never_cache
def descargar_respaldo(request, respaldo_id):
    file_path = os.path.join(settings.BASE_DIR, 'backups', respaldo_id)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/octet-stream")
            response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_path)
            return response
    raise Http404


@method_decorator(never_cache, name='post')
class EliminarRespaldoView(View):
    def post(self, request, *args, **kwargs):
        respaldo_id = request.POST.get('respaldo_id')
        if respaldo_id:
            file_path = os.path.join(settings.BASE_DIR, 'backups', respaldo_id)
            
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    messages.success(request, f"Respaldo {respaldo_id} eliminado exitosamente")
                except Exception as e:
                    messages.error(request, f"Error al eliminar el archivo: {str(e)}")
            else:
                messages.error(request, f"El archivo {respaldo_id} no existe")
        else:
            messages.error(request, "No se especificó un archivo para eliminar")

        return redirect('backup')


@never_cache
def filtrar_respaldos(request):
    respaldos = Respaldo.objects.all()

    nombre_archivo = request.GET.get('nombre_archivo')
    fecha_creacion = request.GET.get('fecha_creacion')
    tamano = request.GET.get('tamano')

    if nombre_archivo:
        respaldos = respaldos.filter(nombre_archivo__icontains=nombre_archivo)
    if fecha_creacion:
        respaldos = respaldos.filter(fecha_creacion__date=fecha_creacion)
    if tamano:
        respaldos = respaldos.filter(tamano__icontains=tamano)

    return render(request, 'backup.html', {'respaldos': respaldos})




#REPORTES CATEGORIA #############################################################################################################
def reporte_categorias_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:C1')
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = "Categorías"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Id", "Nombre", "Estado"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    categorias = Categoria.objects.all()
    for categoria in categorias:
        ws.append([
            categoria.id,
            categoria.nombre,
            'Activado' if categoria.activado else 'Inactivado'
        ])

    column_widths = [10, 30, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_categorias.xlsx"'
    
    wb.save(response)
    return response


#REPORTES CATEGORIA PDF #############################################################################################################
def reporte_categorias_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    table_width = width - 2 * margin
    row_height = 20
    y_position = height - margin

    watermark_path = finders.find('img/logo.png')
    p.saveState()


    p.setFillColor(Color(1, 1, 1, alpha=0.3)) 
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3)) 

    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')

    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Categorías")
    y_position -= 50

    column_widths = [30, 80, 60]
    headers = ["Id", "Nombre", "Estado"]
    data = [headers]

    categorias = Categoria.objects.all()
    for categoria in categorias:
        data.append([
            str(categoria.id),
            categoria.nombre,
            'Activado' if categoria.activado else 'Inactivado'
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)
    
    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height
    
    table.drawOn(p, table_x, table_y) 

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_categorias.pdf"'
    return response

#REPORTES COMPRAS EXCEL #############################################################################################################

def reporte_compras_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:G1') 
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')  
    ws['A2'] = "Compras"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre del producto", "Fecha ingreso", "Cantidad", "Valor unitario", "Valor total", "Estado"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    compras = compra.objects.all()
    for compra_item in compras:
        ws.append([
            compra_item.id,
            compra_item.nombreproducto,
            compra_item.fechaingreso,
            compra_item.cantidad,
            compra_item.valorunitario,
            compra_item.valortotal,
            'Activado' if compra_item.activado else 'Inactivado'
        ])

    column_widths = [10, 30, 20, 10, 15, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_compras.xlsx"'
    
    wb.save(response)
    return response


#REPORTES COMPRAS PDF #############################################################################################################

def reporte_compras_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    table_width = width - 2 * margin
    row_height = 20
    y_position = height - margin

    watermark_path = finders.find('img/logo.png')
    p.saveState()


    p.setFillColor(Color(1, 1, 1, alpha=0.3)) 
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3)) 

    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')

    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Reporte de Compras")
    y_position -= 50

    column_widths = [30, 150, 80, 60, 60, 80, 60]
    headers = ["ID", "Nombre del producto", "Fecha ingreso", "Cantidad", "Valor unitario", "Valor total", "Estado"]
    data = [headers]

    compras = compra.objects.all()
    for compra_obj in compras:
        data.append([
            str(compra_obj.id),
            compra_obj.nombreproducto,
            compra_obj.fechaingreso.strftime("%d/%m/%Y"),
            str(compra_obj.cantidad),
            str(compra_obj.valorunitario),
            str(compra_obj.valortotal),
            'Activado' if compra_obj.activado else 'Inactivado'
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
         ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)
    
    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height
    
    table.drawOn(p, table_x, table_y) 

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_compras.pdf"'
    return response





#REPORTES MARCAS EXCEL #############################################################################################################

def reporte_marcas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marcas"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:C1')
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = "Marcas"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Id", "Nombre", "Estado"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    marcas = Marca.objects.all()
    for marca in marcas:
        ws.append([
            marca.id,
            marca.nombre,
            'Activado' if marca.activado else 'Inactivado'
        ])


    column_widths = [10, 30, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_marcas.xlsx"'
    
    wb.save(response)
    return response

#REPORTES MARCAS PDF #############################################################################################################

def reporte_marcas_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    table_width = width - 2 * margin
    row_height = 20
    y_position = height - margin

    watermark_path = finders.find('img/logo.png')
    p.saveState()

    p.setFillColor(Color(1, 1, 1, alpha=0.3)) 
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3)) 

    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')

    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Marcas")
    y_position -= 50

    column_widths = [30, 80, 60]
    headers = ["Id", "Nombre", "Estado"]
    data = [headers]

    marcas = Marca.objects.all()
    for marca in marcas:
        data.append([
            str(marca.id),
            marca.nombre,
            'Activado' if marca.activado else 'Inactivado'
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)
    
    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height
    
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_marcas.pdf"'
    return response

#REPORTES PROVEEDORES EXCEL #############################################################################################################

def reporte_proveedores_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:F1')
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = "Proveedores"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Id", "Nombre", "Número de Celular", "Correo Electrónico", "Marca perteneciente", "Estado"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    proveedores = Proveedor.objects.all()
    for proveedor in proveedores:
        marca_nombre = proveedor.marca.nombre if proveedor.marca else 'Sin Marca'
        ws.append([
            proveedor.id,
            proveedor.nombre,
            proveedor.numero_celular,
            proveedor.correo_electronico,
            marca_nombre,
            'Activado' if proveedor.activado else 'Inactivado'
        ])

    column_widths = [10, 40, 20, 30, 25, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_proveedores.xlsx"'
    
    wb.save(response)
    return response

#REPORTES PROVEEDORES PDF #############################################################################################################

def reporte_proveedores_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    extra_margin = 20
    row_height = 18

    watermark_path = finders.find('img/logo.png')
    p.saveState()

    p.setFillColor(Color(1, 1, 1, alpha=0.3))
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3))

    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')
    p.restoreState()

    y_position = height - margin
    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Proveedores")
    y_position -= 50

    # Reduce the column widths
    column_widths = [20, 150, 75, 150, 100, 80]
    headers = ["ID", "Nombre", "Número de Celular", "Correo Electrónico", "Marca perteneciente", "Estado"]
    data = [headers]

    proveedores = Proveedor.objects.all()
    for proveedor in proveedores:
        marca_nombre = proveedor.marca.nombre if proveedor.marca else 'Sin Marca'
        data.append([
            str(proveedor.id),
            proveedor.nombre,
            proveedor.numero_celular,
            proveedor.correo_electronico,
            marca_nombre,
            'Activado' if proveedor.activado else 'Inactivado'
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 8),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)

    table_width = sum(column_widths)
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height - extra_margin

    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_proveedores.pdf"'
    return response




#REPORTES PRESENTACIONES EXCEL #############################################################################################################

def reporte_presentaciones_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presentaciones"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:C1')
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = "Presentaciones"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Id", "Nombre", "Estado"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    presentaciones = Presentacion.objects.all()
    for presentacion in presentaciones:
        ws.append([
            presentacion.id,
            presentacion.nombre,
            'Activado' if presentacion.activado else 'Inactivado'
        ])

    column_widths = [10, 30, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_presentaciones.xlsx"'
    
    wb.save(response)
    return response

#REPORTES PRESENTACIONES PDF #############################################################################################################

def reporte_presentaciones_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    table_width = width - 2 * margin
    row_height = 20
    y_position = height - margin

    watermark_path = finders.find('img/logo.png')
    p.saveState()
    p.setFillColor(Color(1, 1, 1, alpha=0.3))  
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3))  
    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')
    p.restoreState()

    
    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Presentaciones")
    y_position -= 50

    
    column_widths = [30, 80, 60]
    headers = ["Id", "Nombre", "Estado"]
    data = [headers]

    presentaciones = Presentacion.objects.all()
    for presentacion in presentaciones:
        data.append([
            str(presentacion.id),
            presentacion.nombre,
            'Activado' if presentacion.activado else 'Inactivado'
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)
    
    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height
    
    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_presentaciones.pdf"'
    return response


##### REPORTE PRODUCTOS EXCEL #########################################


def reporte_productos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    logo_path = finders.find('img/logo.png')
    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:H1')
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = "Productos"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Id", "Nombre", "Precio", "Categoría", "Marca", "Presentación", "Unidad Medida", "Estado"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    productos = Producto.objects.all()
    for producto in productos:
        ws.append([
            producto.id,
            producto.nombre,
            producto.precio,
            str(producto.categoria),
            str(producto.marca),
            str(producto.presentacion),
            str(producto.unidad_medida),
            'Activado' if producto.activado else 'Inactivado'
        ])

    column_widths = [10, 30, 15, 20, 20, 30, 30, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_productos.xlsx"'
    
    wb.save(response)
    return response



#REPORTES PRODUCTOS PDF #############################################################################################################

def reporte_productos_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    table_width = width - 2 * margin
    row_height = 20
    y_position = height - margin

    watermark_path = finders.find('img/logo.png')
    p.saveState()

    p.setFillColor(Color(1, 1, 1, alpha=0.3)) 
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3)) 

    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')

    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Productos")
    y_position -= 50

    column_widths = [30, 80, 50, 60, 60, 80, 80, 60]
    headers = ["Id", "Nombre", "Precio", "Categoría", "Marca", "Presentación", "Unidad Medida", "Estado"]
    data = [headers]

    productos = Producto.objects.all()
    for producto in productos:
        data.append([
            str(producto.id),
            producto.nombre,
            str(producto.precio),
            str(producto.categoria),
            str(producto.marca),
            str(producto.presentacion),
            str(producto.unidad_medida),
            'Activado' if producto.activado else 'Inactivado'
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)
    
    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_productos.pdf"'
    return response

############ REPORTE UNIDAD DE MEDIDA EXCEL #################################

def reporte_unidades_medida_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unidades de medida"

    logo_path = finders.find('img/logo.png')
    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:C1')
    ws['B1'] = "TIENDA BONANZA"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = "Unidades de medida"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Id", "Nombre", "Estado"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    unidades_medida = UnidadMedida.objects.all()
    for unidad in unidades_medida:
        ws.append([
            unidad.id,
            unidad.nombre,
            'Activado' if unidad.activado else 'Inactivado'
        ])

    column_widths = [10, 30, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_unidades_medida.xlsx"'
    
    wb.save(response)
    return response


#REPORTES UNIDADES PDF #############################################################################################################

def reporte_unidades_medida_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    table_width = width - 2 * margin
    row_height = 20
    y_position = height - margin

    watermark_path = finders.find('img/logo.png')
    p.saveState()
    p.setFillColor(Color(1, 1, 1, alpha=0.3)) 
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3)) 
    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')
    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Unidades de Medida")
    y_position -= 50

    column_widths = [30, 80, 60]
    headers = ["Id", "Nombre", "Estado"]
    data = [headers]

    unidades_medida = UnidadMedida.objects.all()
    for unidad in unidades_medida:
        data.append([
            str(unidad.id),
            unidad.nombre,
            'Activado' if unidad.activado else 'Inactivado'
        ])

    table = Table(data, colWidths=column_widths)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (0, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])
    table.setStyle(style)
    
    table.wrapOn(p, width - 2 * margin, height - 2 * margin)
    
    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height
    
    table.drawOn(p, table_x, table_y) 

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_unidades_medida.pdf"'
    return response


#REPORTES VENTAS EXCEL #############################################################################################################

def reporte_ventas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 65
    img.width = 85
    ws.add_image(img, 'A1')

    ws.merge_cells('A2:G2')
    ws['A2'] = "TIENDA BONANZA"
    ws['A2'].font = Font(size=24, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:G3')
    ws['A3'] = "Ventas"
    ws['A3'].font = Font(size=18)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre del producto", "Fecha venta", "Cantidad", "Valor unitario", "Valor total", "Estado"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="25b6e6", end_color="25b6e6", fill_type="solid")
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ventas = venta.objects.all()
    for venta_item in ventas:
        ws.append([
            venta_item.id,
            venta_item.nombreproducto,
            venta_item.fechaventa,
            venta_item.cantidad,
            venta_item.valorunitario,
            venta_item.valortotal,
            'Activado' if venta_item.activado else 'Inactivado'
        ])

    column_widths = [10, 25, 15, 10, 15, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    for cell in ws[4]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_ventas.xlsx"'
    
    wb.save(response)
    return response


#REPORTES VENTAS PDF #############################################################################################################

def reporte_ventas_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margin = 50
    table_width = width - 2 * margin
    row_height = 20
    y_position = height - margin

    watermark_path = finders.find('img/logo.png')
    p.saveState()
    p.setFillColor(Color(1, 1, 1, alpha=0.3))
    p.setStrokeColor(Color(1, 1, 1, alpha=0.3))
    p.drawImage(watermark_path, x=(width - 400) / 2, y=(height - 400) / 2, width=400, height=400, mask='auto')
    p.restoreState()

    p.setFont("Helvetica-Bold", 24)
    p.drawString(margin, y_position, "TIENDA BONANZA")
    y_position -= 30
    p.setFont("Helvetica", 18)
    p.drawString(margin, y_position, "Ventas")
    y_position -= 50

    column_widths = [30, 100, 80, 60, 60, 80, 60]
    headers = ["ID", "Nombre del producto", "Fecha venta", "Cantidad", "Valor unitario", "Valor total", "Estado"]
    data = [headers]

    ventas_list = venta.objects.all()
    for v in ventas_list:
        data.append([
            str(v.id),
            v.nombreproducto,
            v.fechaventa.strftime('%Y-%m-%d'),
            str(v.cantidad),
            str(v.valorunitario),
            str(v.valortotal),
            'Activado' if v.activado else 'Inactivado'
        ])

    table = Table(data, colWidths=column_widths)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#25b6e6'),
        ('TEXTCOLOR', (0, 0), (-1, 0), (0, 0, 0)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0)),
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 1), (-1, -1), '#f2f2f2'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER')
    ])

    table.setStyle(style)

    table.wrapOn(p, width - 2 * margin, height - 2 * margin)

    table_x = (width - table_width) / 2
    table_y = y_position - len(data) * row_height

    table.drawOn(p, table_x, table_y)

    p.showPage()
    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_ventas.pdf"'
    return response


#FILTROS Y BUSQUEDA##############################################################################################################

#FILTROS CATEGORIA 

def categoria_list(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')

    if estado_filtro == 'activado':
        categorias = Categoria.objects.filter(activado=True)
    elif estado_filtro == 'inactivado':
        categorias = Categoria.objects.filter(activado=False)
    else:
        categorias = Categoria.objects.all()

    if buscar:
        categorias = categorias.filter(nombre__icontains=buscar)
    
    context = {
        'categorias': categorias
    }
    
    return render(request, 'categorias/index.html', context)



#FILTROS MARCA###############################################################################################################

def marca_list(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')

    if estado_filtro == 'activado':
        marcas = Marca.objects.filter(activado=True)    
    elif estado_filtro == 'inactivado':
        marcas = Marca.objects.filter(activado=False)
    else:
        marcas = Marca.objects.all()

    if buscar:
        marcas = marcas.filter(nombre__icontains=buscar)
    
    context = {
        'marcas': marcas
    }
    
    return render(request, 'marcas/index.html', context)


#FILTROS PRESENTACION ###############################################################################################################

def presentacion_list(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')

    if estado_filtro == 'activado':
        presentaciones = Presentacion.objects.filter(activado=True)    
    elif estado_filtro == 'inactivado':
        presentaciones = Presentacion.objects.filter(activado=False)
    else:
        presentaciones = Presentacion.objects.all()

    if buscar:
        presentaciones = presentaciones.filter(nombre__icontains=buscar)
    
    context = {
        'presentaciones': presentaciones
    }
    
    return render(request, 'presentaciones/index.html', context)


#FILTROSUNIDAD DE MEDIDA ###############################################################################################################

def unidades_medida_list(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')

    if estado_filtro == 'activado':
        unidades_medida = UnidadMedida.objects.filter(activado=True)    
    elif estado_filtro == 'inactivado':
        unidades_medida = UnidadMedida.objects.filter(activado=False)
    else:
        unidades_medida = UnidadMedida.objects.all()

    if buscar:
        unidades_medida = unidades_medida.filter(nombre__icontains=buscar)
    
    context = {
        'unidades_medida': unidades_medida
    }
    
    return render(request, 'unidades_medida/index.html', context)

#FILTROS PRODUCTO ###############################################################################################################

def producto_list(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')
    precio_min = request.GET.get('precio_min', None)
    precio_max = request.GET.get('precio_max', None)
    categoria_id = request.GET.get('categoria', '')
    marca_id = request.GET.get('marca', '')
    presentacion_id = request.GET.get('presentacion', '')
    unidad_medida_id = request.GET.get('unidad_medida', '')

    productos = Producto.objects.all()

    if estado_filtro == 'activado':
        productos = productos.filter(activado=True)
    elif estado_filtro == 'inactivado':
        productos = productos.filter(activado=False)

    if buscar:
        productos = productos.filter(nombre__icontains=buscar)
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    if marca_id:
        productos = productos.filter(marca_id=marca_id)
    if presentacion_id:
        productos = productos.filter(presentacion_id=presentacion_id)
    if unidad_medida_id:
        productos = productos.filter(unidad_medida_id=unidad_medida_id)

    categorias = Categoria.objects.all()
    marcas = Marca.objects.all()
    presentaciones = Presentacion.objects.all()
    unidades_medida = UnidadMedida.objects.all()

    context = {
        'productos': productos,
        'categorias': categorias,
        'marcas': marcas,
        'presentaciones': presentaciones,
        'unidades_medida': unidades_medida
    }

    return render(request, 'producto/index.html', context)

#FILTROS VENTA ###############################################################################################################

def venta_list(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')
    fecha_filtro = request.GET.get('fecha', None)
    valor_unitario_filtro = request.GET.get('valorunitario', None)
    valor_total_filtro = request.GET.get('valortotal', None)

    ventas = venta.objects.all()

    if estado_filtro == 'activado':
        ventas = ventas.filter(activado=True)
    elif estado_filtro == 'inactivado':
        ventas = ventas.filter(activado=False)

    if buscar:
        ventas = ventas.filter(nombreproducto__icontains=buscar)

    if fecha_filtro:
        ventas = ventas.filter(fechaventa=fecha_filtro)

    if valor_unitario_filtro:
        ventas = ventas.filter(valorunitario=valor_unitario_filtro)

    if valor_total_filtro:
        ventas = ventas.filter(valortotal=valor_total_filtro)

    context = {
        'ventas': ventas
    }

    return render(request, 'venta/index.html', context)

#FILTROS COMPRA ###############################################################################################################

def compra_list(request):
    estado_filtro = request.GET.get('estado', None)
    buscar = request.GET.get('buscar', '')
    fecha_filtro = request.GET.get('fecha', None)
    valor_unitario_filtro = request.GET.get('valorunitario', None)
    valor_total_filtro = request.GET.get('valortotal', None)

    compras = compra.objects.all()

    if estado_filtro == 'activado':
        compras = compras.filter(activado=True)
    elif estado_filtro == 'inactivado':
        compras = compras.filter(activado=False)

    if buscar:
        compras = compras.filter(nombreproducto__icontains=buscar)

    if fecha_filtro:
        compras = compras.filter(fechaingreso=fecha_filtro)

    if valor_unitario_filtro:
        compras = compras.filter(valorunitario=valor_unitario_filtro)

    if valor_total_filtro:
        compras = compras.filter(valortotal=valor_total_filtro)

    context = {
        'compras': compras
    }

    return render(request, 'compra/index.html', context)


#FILTROS PROVEEDOR ###############################################################################################################

def proveedor_list(request):
    estado_filtro = request.GET.get('estado', None)
    buscar_nombre = request.GET.get('buscar_nombre', '')
    buscar_celular = request.GET.get('buscar_celular', '')
    buscar_correo = request.GET.get('buscar_correo', '')
    marca_filtro = request.GET.get('marca', None)

    proveedores = Proveedor.objects.all()

    if estado_filtro == 'activado':
        proveedores = proveedores.filter(activado=True)
    elif estado_filtro == 'inactivado':
        proveedores = proveedores.filter(activado=False)

    if buscar_nombre:
        proveedores = proveedores.filter(nombre__icontains=buscar_nombre)

    if buscar_celular:
        proveedores = proveedores.filter(numero_celular__icontains=buscar_celular)

    if buscar_correo:
        proveedores = proveedores.filter(correo_electronico__icontains=buscar_correo)

    if marca_filtro:
        proveedores = proveedores.filter(marca_id=marca_filtro)

    context = {
        'proveedores': proveedores,
        'estado_filtro': estado_filtro,
        'buscar_nombre': buscar_nombre,
        'buscar_celular': buscar_celular,
        'buscar_correo': buscar_correo,
        'marcas': Marca.objects.all(),  
    }

    return render(request, 'proveedor/index.html', context)




#ESTADOS ###############################################################

def activar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    categoria.activado = True
    categoria.save()
    return redirect('categorias')

def desactivar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    categoria.activado = False
    categoria.save()
    return redirect('categorias')

def activar_marca(request, pk):
    marca = get_object_or_404(Marca, pk=pk)
    marca.activado = True
    marca.save()
    return redirect('marcas')

def desactivar_marca(request, pk):
    presentacion = get_object_or_404(Marca, pk=pk)
    presentacion.activado = False
    presentacion.save()
    return redirect('marcas')

def activar_presentacion(request, pk):
    presentacion = get_object_or_404(Presentacion, pk=pk)
    presentacion.activado = True
    presentacion.save()
    return redirect('presentaciones')

def desactivar_presentacion(request, pk):
    marca = get_object_or_404(Presentacion, pk=pk)
    marca.activado = False
    marca.save()
    return redirect('presentaciones')

def activar_unidades_medida(request, pk):
    unidades_medida = get_object_or_404(UnidadMedida, pk=pk)
    unidades_medida.activado = True
    unidades_medida.save()
    return redirect('unidades_medida')

def desactivar_unidades_medida(request, pk):
    unidades_medida = get_object_or_404(UnidadMedida, pk=pk)
    unidades_medida.activado = False
    unidades_medida.save()
    return redirect('unidades_medida')

def activar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    producto.activado = True
    producto.save()
    return redirect('producto')

def desactivar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    producto.activado = False
    producto.save()
    return redirect('producto')

def activar_compra(request, pk):
    compras = get_object_or_404(compra, pk=pk)
    compras.activado = True
    compras.save()
    return redirect('compra')

def desactivar_compra(request, pk):
    compras = get_object_or_404(compra, pk=pk)
    compras.activado = False
    compras.save()
    return redirect('compra')

def activar_venta(request, pk):
    ventas = get_object_or_404(venta, pk=pk)
    ventas.activado = True
    ventas.save()
    return redirect('venta')

def desactivar_venta(request, pk):
    ventas = get_object_or_404(venta, pk=pk)
    ventas.activado = False
    ventas.save()
    return redirect('venta')

def activar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    proveedor.activado = True
    proveedor.save()
    return redirect('proveedores')

def desactivar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    proveedor.activado = False
    proveedor.save()
    return redirect('proveedores')




#CATEGORIAS################################################################################################################


@never_cache
@login_required
def categorias(request):
    categorias = Categoria.objects.all()
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Categorías', 'url': '/libreria/categorias/'},
    ]
    return render(request, 'categorias/index.html', {'categorias': categorias, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def crear_categorias(request):
    form = CategoriaForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, '¡Categoria creada con éxito!')
        return redirect('categorias')
    breadcrumbs = [

        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Categorías', 'url': '/libreria/categorias/'},
        {'name': 'Crear Categorías', 'url': '/libreria/categorias/crear_categorias/'}, 
    ]

    return render(request, 'categorias/crear.html', {'form': form, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def editar_categorias(request, pk):
    categoria = Categoria.objects.get(pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, request.FILES or None, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Categoría editada con éxito!')
            return redirect('categorias')
    else:
        form = CategoriaForm(instance=categoria)

        breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Categorías', 'url': '/libreria/categorias/'},
        {'name': 'Editar Categorías', 'url': f'/libreria/categorias/editar_categorias/{pk}/'}, 
    ]
    return render(request, 'categorias/editar.html', {'form': form, 'categoria': categoria, 'breadcrumbs': breadcrumbs})



@never_cache
@login_required
def eliminar_categorias(request, pk):
    categoria = Categoria.objects.get(pk=pk)
    categoria.delete()
    messages.success(request, '¡Categoría eliminada con éxito!')
    return redirect('categorias')


#MARCAS ########################################################################################################################

@never_cache
@login_required
def marcas(request):
    marcas = Marca.objects.all()
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Marcas', 'url': '/libreria/marcas/'},
    ]
    return render(request, 'marcas/index.html', {'marcas': marcas,'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def crear_marcas(request):
    form = MarcaForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, '¡Marca creada con éxito!')
        return redirect('marcas')
    breadcrumbs = [

        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Marcas', 'url': '/libreria/marcas/'},
        {'name': 'Crear Marcas', 'url': '/libreria/marcas/crear_marcas/'},  
    ]

    return render(request, 'marcas/crear.html', {'form': form, 'breadcrumbs': breadcrumbs})



@never_cache
@login_required
def editar_marcas(request, pk):
    marca = Marca.objects.get(pk=pk)
    if request.method == 'POST':
        form = MarcaForm(request.POST, request.FILES or None, instance=marca)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Marca editada con éxito!')
            return redirect('marcas')
    else:
        form = MarcaForm(instance=marca)
        breadcrumbs = [

        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Marcas', 'url': '/libreria/marcas/'},
        {'name': 'Editar Marcas', 'url': f'/libreria/marcas/editar_marcas/{pk}/'}, 
    ]

    return render(request, 'marcas/editar.html', {'form': form, 'marca': marca,  'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def eliminar_marcas(request, pk):
    marca = Marca.objects.get(pk=pk)
    marca.delete()
    messages.success(request, '¡Marca eliminada con éxito!')
    return redirect('marcas')

#PRESENTACIONES #####################################################################################################

@never_cache
@login_required
def presentaciones(request):
    presentaciones = Presentacion.objects.all()
    breadcrumbs = [

        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Presentaciones', 'url': '/libreria/presentaciones/'},
    
    ]
    return render(request, 'presentaciones/index.html', {'presentaciones': presentaciones, 'breadcrumbs': breadcrumbs})



@never_cache
@login_required
def crear_presentaciones(request):
    form = PresentacionForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, '¡Presentación creada con éxito!')
        return redirect('presentaciones')
    breadcrumbs = [

        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Presentaciones', 'url': '/libreria/presentaciones/'},
        {'name': 'Crear Presentaciones', 'url': '/libreria/presentaciones/crear_presentaciones/'},  
    ]
    return render(request, 'presentaciones/crear.html', {'form': form, 'breadcrumbs': breadcrumbs})



@never_cache
@login_required
def editar_presentaciones(request, pk):
    presentacion = Presentacion.objects.get(pk=pk)
    if request.method == 'POST':
        form = PresentacionForm(request.POST, request.FILES or None, instance=presentacion)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Presentación editada con éxito!')
            return redirect('presentaciones')
    else:
        form = PresentacionForm(instance=presentacion)
        breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Presentaciones', 'url': '/libreria/presentaciones/'},
        {'name': 'Editar Presentaciones', 'url': f'/libreria/presentaciones/editar_presentaciones/{pk}/'}, 
    ]

    return render(request, 'presentaciones/editar.html', {'form': form, 'presentacion': presentacion, 'breadcrumbs': breadcrumbs})




@never_cache
@login_required
def eliminar_presentaciones(request, pk):
    presentacion = Presentacion.objects.get(pk=pk)
    presentacion.delete()
    messages.success(request, '¡Presentación eliminada con éxito!')
    return redirect('presentaciones')


#UNIDAD DE MEDIDA  #####################################################################################################

@never_cache
@login_required
def unidades_medida(request):
    unidades_medida = UnidadMedida.objects.all()
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Unidades de Medida', 'url': '/libreria/unidades_medida/'},
    ]

    return render(request, 'unidades_medida/index.html', {'unidades_medida': unidades_medida, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def crear_unidades_medida(request):
    form = UnidadMedidaForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, '¡Unidad de medida creada con éxito!')
        return redirect('unidades_medida')
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Unidades de medida', 'url': '/libreria/unidades_medida/'},
        {'name': 'Crear Unidad de medida', 'url': '/libreria/unidades_medida/crear_unidades_medida/'},  
    ]
    return render(request, 'unidades_medida/crear.html', {'form': form, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def editar_unidades_medida(request, pk):
    unidad_medida = UnidadMedida.objects.get(pk=pk)
    if request.method == 'POST':
        form = UnidadMedidaForm(request.POST, request.FILES or None, instance=unidad_medida)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Unidad de medida editada con éxito!')
            return redirect('unidades_medida')
    else:
        form = UnidadMedidaForm(instance=unidad_medida)
        
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Unidades de medida', 'url': '/libreria/unidades_medida/'},
        {'name': 'Editar Unidades de medida', 'url': f'/libreria/unidades_medida/editar_unidades_medida/{pk}/'},
          
    ]    
    return render(request, 'unidades_medida/editar.html', {'form': form, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def eliminar_unidades_medida(request, pk):
    unidad_medida = UnidadMedida.objects.get(pk=pk)
    unidad_medida.delete()
    messages.success(request, '¡Unidad de medida eliminada con éxito!')
    return redirect('unidades_medida')

#PRODUCTO #####################################################################################################

@never_cache
@login_required
def producto(request):
    productos = Producto.objects.all()
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Productos', 'url': '/libreria/producto/'},
    ]

    return render(request, 'producto/index.html', {'productos': productos, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def crear_productos(request):
    form = ProductoForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, '¡Producto agregado con éxito!')
        return redirect('producto')
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Productos', 'url': '/libreria/producto/'},
        {'name': 'Crear Productos', 'url': '/libreria/producto/crear_productos/'},  
    ]
    
    return render(request, 'producto/crear.html', {'form': form, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def editar_productos(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES or None, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Producto editado con éxito!')
            return redirect('producto')
    else:
        form = ProductoForm(instance=producto)
    
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Productos', 'url': '/libreria/producto/'},
        {'name': 'Editar Producto', 'url': f'/libreria/producto/editar_productos/{pk}/'},  
    ]
    return render(request, 'producto/editar.html', {'form': form, 'producto': producto, 'breadcrumbs': breadcrumbs})
    


@never_cache
@login_required
def eliminar_productos(request, pk):
    producto = Producto.objects.get(pk=pk)
    producto.delete()
    messages.success(request, '¡Producto eliminado con éxito!')
    return redirect('producto')


# proveedores #####################################################################################################

@never_cache
@login_required
def proveedores(request):
    proveedores = Proveedor.objects.all()

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Proveedores', 'url': '/libreria/proveedores/'},
    ]
    return render(request, 'proveedor/index.html', {'proveedores': proveedores, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        formulario = ProveedorForm(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, '¡Proveedor creado con éxito!')
            return redirect('proveedores')
    else:
        formulario = ProveedorForm()
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Proveedores', 'url': '/libreria/proveedores/'},
        {'name': 'Crear Proveedores', 'url': '/libreria/proveedores/crear/'},
    ]
        
    return render(request, 'proveedor/crear.html', {'formulario': formulario, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def editar_proveedor(request, id):
    proveedor_ed = Proveedor.objects.get(id=id)
    if request.method == 'POST':
        formulario = ProveedorForm(request.POST, instance=proveedor_ed)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, '¡Proveedor editado con éxito!')
            return redirect('proveedores')
    else:
        formulario = ProveedorForm(instance=proveedor_ed)
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Proveedores', 'url': '/libreria/proveedores/'},
        {'name': 'Editar Proveedores', 'url': f'/libreria/proveedores/editar/{id}/'},  
    ]
    return render(request, 'proveedor/editar.html', {'formulario': formulario, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def eliminar_proveedor(request, id):
    proveedor = Proveedor.objects.get(id=id)
    proveedor.delete()
    messages.success(request, '¡Proveedor eliminado con éxito!')
    return redirect('proveedores')


#COMPRA #####################################################################################################

@never_cache
@login_required
def crear_compra(request):
    productos = Producto.objects.all()
    if request.method == 'POST':
        formulario = CompraForm(request.POST)
        if formulario.is_valid():
            try:
                formulario.save()
                messages.success(request, '¡Compra creada con éxito!')
                return redirect('compra')
            except Exception as e:
                messages.error(request, f'Error al crear la compra: {e}')
        else:
            # Print form errors for debugging
            print(formulario.errors)
    else:
        formulario = CompraForm()
    
    return render(request, 'compra/crear.html', {'formulario': formulario, 'productos': productos})

@never_cache
@login_required
def editar_compra(request, id):
    compra_ed = get_object_or_404(Compra, id=id)
    if request.method == 'POST':
        formulario = CompraForm(request.POST, instance=compra_ed)
        if formulario.is_valid():
            try:
                formulario.save()
                messages.success(request, '¡Compra editada con éxito!')
                return redirect('compra')
            except Exception as e:
                messages.error(request, f'Error al editar la compra: {e}')
        else:
            # Print form errors for debugging
            print(formulario.errors)
    else:
        formulario = CompraForm(instance=compra_ed)

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Compras', 'url': '/libreria/compra/'},
        {'name': 'Editar Compra', 'url': f'/libreria/compra/editar/{id}/'},  
    ]
    return render(request, 'compra/editar.html', {'formulario': formulario, 'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def eliminar_compra(request, id):
    comprae = get_object_or_404(Compra, id=id)
    comprae.delete()
    messages.success(request, '¡Compra eliminada con éxito!')
    return redirect('compra')

#VENTA #####################################################################################################


@never_cache
@login_required
def ventas(request):
    ventas = venta.objects.all()
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Ventas', 'url': '/libreria/venta/'},
    ]
    return render(request, 'venta/index.html', {'ventas': ventas, 'breadcrumbs': breadcrumbs})



@never_cache
@login_required
def crear_venta(request):
    if request.method == 'POST':
        formulario = VentaForm(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, '¡Venta creada con éxito!')
            return redirect('venta')
    else:
        formulario = VentaForm()
    
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Ventas', 'url': '/libreria/venta/'},
        {'name': 'Crear Venta', 'url': '/libreria/venta/crear'},  
    ]
    
    return render(request, 'venta/crear.html', {'formulario': formulario, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def editar_venta(request, id):
    venta_ed = get_object_or_404(venta, id=id)
    if request.method == 'POST':
        formulario = VentaForm(request.POST, instance=venta_ed)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, '¡Venta editada con éxito!')
            return redirect('venta')
    else:
        formulario = VentaForm(instance=venta_ed)
    
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
        {'name': 'Ventas', 'url': '/libreria/venta/'},
        {'name': 'Editar Venta', 'url': f'/libreria/venta/editar/{id}/'},  
    ]
    
    return render(request, 'venta/editar.html', {'formulario': formulario, 'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def eliminar_venta(request, id):
    ventae = venta.objects.get(id=id)
    ventae.delete()
    messages.success(request, '¡Venta eliminada con éxito!')
    return redirect('venta')

@never_cache
def inicio(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
    ]
    return render(request, "inicio.html", {'breadcrumbs': breadcrumbs})

@never_cache
def respaldos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/dashboard'},
    ]
    return render(request, "backup.html", {'breadcrumbs': breadcrumbs})

@never_cache
@login_required
def registrarme (request):
   return render(request, "registrarme.html")


@never_cache
@login_required
def dashboard (request):
    breadcrumbs = [
        {'name': 'Inicio/', 'url': '/dashboard'},
    ]
    return render(request, "dashboard.html",  {'breadcrumbs': breadcrumbs})


@never_cache
@login_required
def empleado (request):
   return render(request, "/empleado")




class loginFormView(LoginView):
    template_name = "login.html"
    
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('dashboard')  
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Iniciar sesión'
        return context

class logoutRedirect(RedirectView):
    pattern_name = 'login'
    
    def dispatch(self, request, *args, **kwargs):
        logout(request)
        return super().dispatch(request, *args, **kwargs)
    
@never_cache
@login_required
def ayuda(request):
    return render(request, 'ayuda.html')








